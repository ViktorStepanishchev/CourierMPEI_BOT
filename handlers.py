from aiogram.types import CallbackQuery, Message
from aiogram import F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
import openpyxl

import keyboard, text
import back
from states import all_state

router = Router()
num_of_order_in_mass = 0
book = openpyxl.open("example.xlsx")
sheet = book.active

@router.message(Command("start"))
async def start(message:Message):
    await message.answer(text.start_text, reply_markup=keyboard.main_keyboard)

@router.callback_query(F.data == 'order')
async def order(clbck: CallbackQuery, state: FSMContext):
    await state.set_state(all_state.wait_message_order)
    await clbck.message.edit_text(text.order_text, reply_markup=keyboard.back_keyboard)

@router.message(all_state.wait_message_order)
async def get_order(message: Message, state: FSMContext):
    await state.update_data(wait_message_order = message.text)
    # back.num_of_order_in_mass += 1
    # back.orders.append([str(back.num_of_order_in_mass), message.text, message.from_user.id])
    # keyboard.create(back.orders)
    back.num_of_order_in_mass += 1
    for row in range(1, 1000):
        if sheet.cell(row=row, column=1).value is None:
            sheet.cell(row=row, column=1).value = back.num_of_order_in_mass
            sheet.cell(row=row, column=2).value = message.text
            sheet.cell(row=row, column=3).value = message.from_user.id
            break
    book.save("example.xlsx")

    await message.answer(text.get_order_text, reply_markup=keyboard.back_keyboard)
    await state.clear()

@router.callback_query(F.data == 'courier')
async def courier(clbck: CallbackQuery):
    keyboard.create()
    await clbck.message.edit_text(text.courier_text, reply_markup=keyboard.order_keyboard.as_markup())

@router.callback_query(F.data == 'help')
async def help(clbck: CallbackQuery):
    await clbck.message.edit_text(text.help_text, reply_markup=keyboard.help_keyboard)

@router.callback_query(F.data == 'back')
async def help(clbck: CallbackQuery):
    await clbck.message.edit_text(text.start_text, reply_markup=keyboard.main_keyboard)

@router.message()
async def start(message:Message):
    await message.answer(text.start_text, reply_markup=keyboard.main_keyboard)